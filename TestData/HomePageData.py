import openpyxl

class HomePageData:

    test_HomePage_Data = [
        {"EmailID" : "supraja@gmail.com", "Pwd" : "welcome_123", "Name" : "Hello", "DataBind": "Hello again"},
        {"EmailID" : "Tanuz@gmail.com", "Pwd" : "welcome_123", "Name" : "Tanuz", "DataBind": "Hello Smile"},
        {"EmailID" : "Hanvi@gmail.com", "Pwd" : "welcome_123", "Name" : "Hanvi", "DataBind": "Hello Hanvi"},
        {"EmailID" : "Sravan@gmail.com", "Pwd" : "welcome_123", "Name" : "Sravan", "DataBind": "Hello Sravan"}]

    @staticmethod
    def get_testdata(test_case_name):
        dic_user = {}
        book = openpyxl.load_workbook("C:\\Users\\supraja.kollapudi\\PycharmProjects\\exceldemo1.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    dic_user[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value  # capture data in dictionary
                    # print(sheet.cell(row=i, column=j).value)
        return[dic_user]
        print("For Git Demo2")
